"""List all BigQuery tables and GCS models, save to Excel"""
import os
import sys
from datetime import datetime

# Add project root to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from google.cloud import bigquery, storage
from ml_logic.secrets import get_secret

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def list_bq_tables():
    """List all tables in the grisou BigQuery dataset."""
    project = get_secret("GCP_PROJECT")
    dataset = get_secret("BQ_DATASET")
    region = get_secret("BQ_REGION")

    client = bigquery.Client(project=project, location=region)
    tables = list(client.list_tables(f"{project}.{dataset}"))

    results = []
    for table in tables:
        table_ref = client.get_table(table.reference)
        results.append({
            "table_name": table.table_id,
            "type": _classify_table(table.table_id),
            "num_rows": table_ref.num_rows,
            "size_mb": round(table_ref.num_bytes / (1024 * 1024), 2) if table_ref.num_bytes else 0,
            "created": table_ref.created.strftime("%Y-%m-%d %H:%M:%S") if table_ref.created else "",
            "modified": table_ref.modified.strftime("%Y-%m-%d %H:%M:%S") if table_ref.modified else "",
        })

    results.sort(key=lambda x: x["created"], reverse=True)
    return results


def list_gcs_models():
    """List all model files in the GCS bucket."""
    bucket_name = get_secret("BUCKET_NAME")

    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blobs = list(bucket.list_blobs(prefix="models/"))

    results = []
    for blob in blobs:
        if blob.name == "models/":
            continue
        results.append({
            "model_name": blob.name.replace("models/", ""),
            "size_mb": round(blob.size / (1024 * 1024), 2) if blob.size else 0,
            "created": blob.time_created.strftime("%Y-%m-%d %H:%M:%S") if blob.time_created else "",
            "updated": blob.updated.strftime("%Y-%m-%d %H:%M:%S") if blob.updated else "",
            "path": f"gs://{bucket_name}/{blob.name}",
        })

    results.sort(key=lambda x: x["created"], reverse=True)
    return results


def _classify_table(name):
    """Classify table by naming convention."""
    if name.startswith("preprocess_"):
        return "preprocessing"
    elif name.startswith("history_"):
        return "training_history"
    elif name.startswith("predictions_"):
        return "predictions"
    else:
        return "raw_data"


def save_to_excel(bq_tables, gcs_models, output_path):
    """Save results to formatted Excel file."""
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Sheet 1: BigQuery Tables ---
    ws_bq = wb.active
    ws_bq.title = "BigQuery Tables"

    bq_headers = ["Table Name", "Type", "Rows", "Size (MB)", "Created", "Last Modified"]
    for col, header in enumerate(bq_headers, 1):
        cell = ws_bq.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, table in enumerate(bq_tables, 2):
        values = [table["table_name"], table["type"], table["num_rows"],
                  table["size_mb"], table["created"], table["modified"]]
        for col, val in enumerate(values, 1):
            cell = ws_bq.cell(row=row_idx, column=col, value=val)
            cell.border = border

    for col in range(1, 7):
        ws_bq.column_dimensions[chr(64 + col)].width = 25

    # --- Sheet 2: GCS Models ---
    ws_gcs = wb.create_sheet("GCS Models")

    gcs_headers = ["Model Name", "Size (MB)", "Created", "Last Updated", "GCS Path"]
    for col, header in enumerate(gcs_headers, 1):
        cell = ws_gcs.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, model in enumerate(gcs_models, 2):
        values = [model["model_name"], model["size_mb"], model["created"],
                  model["updated"], model["path"]]
        for col, val in enumerate(values, 1):
            cell = ws_gcs.cell(row=row_idx, column=col, value=val)
            cell.border = border

    for col in range(1, 6):
        ws_gcs.column_dimensions[chr(64 + col)].width = 35

    # --- Sheet 3: Summary ---
    ws_summary = wb.create_sheet("Summary")
    ws_summary.insert_rows(0)

    summary_data = [
        ["Grisounet Resources Summary", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["BigQuery", ""],
        ["Total tables", len(bq_tables)],
        ["Raw data tables", sum(1 for t in bq_tables if t["type"] == "raw_data")],
        ["Preprocessing tables", sum(1 for t in bq_tables if t["type"] == "preprocessing")],
        ["History tables", sum(1 for t in bq_tables if t["type"] == "training_history")],
        ["Prediction tables", sum(1 for t in bq_tables if t["type"] == "predictions")],
        ["", ""],
        ["Cloud Storage", ""],
        ["Total models", len(gcs_models)],
        ["Total size (MB)", round(sum(m["size_mb"] for m in gcs_models), 2)],
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws_summary.cell(row=row_idx, column=1, value=label)
        cell_b = ws_summary.cell(row=row_idx, column=2, value=value)
        if label in ["Grisounet Resources Summary", "BigQuery", "Cloud Storage"]:
            cell_a.font = Font(bold=True, size=12)

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 25

    wb.save(output_path)


def save_to_txt(bq_tables, gcs_models, output_path):
    """Save results to plain text file."""
    lines = []
    lines.append(f"GRISOUNET RESOURCES — Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 70)

    lines.append(f"\nBIGQUERY TABLES ({len(bq_tables)} total)")
    lines.append("-" * 70)
    for t in bq_tables:
        lines.append(f"  {t['table_name']:40s} | {t['type']:18s} | {t['num_rows']:>8} rows | {t['created']}")

    lines.append(f"\nGCS MODELS ({len(gcs_models)} total)")
    lines.append("-" * 70)
    for m in gcs_models:
        lines.append(f"  {m['model_name']:40s} | {m['size_mb']:>8.2f} MB | {m['created']}")
        lines.append(f"    {m['path']}")

    with open(output_path, "w") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    print("Fetching BigQuery tables...")
    bq_tables = list_bq_tables()
    print(f"  Found {len(bq_tables)} tables")

    print("Fetching GCS models...")
    gcs_models = list_gcs_models()
    print(f"  Found {len(gcs_models)} models")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs("results/inventory", exist_ok=True)

    if HAS_OPENPYXL:
        output = f"results/inventory/resources_{timestamp}.xlsx"
        save_to_excel(bq_tables, gcs_models, output)
    else:
        output = f"results/inventory/resources_{timestamp}.txt"
        save_to_txt(bq_tables, gcs_models, output)
        print("  (Install openpyxl for Excel output: pip install openpyxl)")

    print(f"\nSaved to {output}")
